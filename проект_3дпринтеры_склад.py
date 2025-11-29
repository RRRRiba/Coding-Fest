import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

COLUMNS = [
    "Дата", "Вид материала", "Размер катушки", "Вес (кг)",
    "Сечение", "Цвет", "Условия хранения", "Статус", "Остаток"
]

DATE_FORMAT = "%d.%m.%Y"


class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Учёт склада (упрощённая версия)")
        self.filepath = None
        self.wb = None
        self.sheet = None
        self.data = []

        self.build_ui()

    def build_ui(self):
        top = ttk.Frame(self.root)
        top.pack(fill="x", padx=5, pady=5)

        ttk.Button(top, text="Открыть", command=self.open_file).pack(side="left", padx=3)
        ttk.Button(top, text="Создать", command=self.create_new).pack(side="left", padx=3)
        ttk.Button(top, text="Сохранить", command=self.save_file).pack(side="left", padx=3)
        ttk.Button(top, text="Сохранить как", command=self.save_as).pack(side="left", padx=3)

        ttk.Label(top, text="Фильтр (по любому полю):").pack(side="left", padx=6)
        self.filter_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.filter_var, width=20).pack(side="left")
        ttk.Button(top, text="Применить", command=self.apply_filter).pack(side="left", padx=3)
        ttk.Button(top, text="Сброс", command=self.reset_filter).pack(side="left", padx=3)

        frame = ttk.Frame(self.root)
        frame.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(frame, columns=COLUMNS, show="headings")
        for col in COLUMNS:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120)
        self.tree.pack(fill="both", expand=True, pady=5)

        form = ttk.Frame(self.root)
        form.pack(fill="x", padx=5, pady=5)

        self.vars = {}
        for i, col in enumerate(COLUMNS):
            ttk.Label(form, text=col).grid(row=i//2, column=(i%2)*2, sticky="w", padx=3, pady=2)
            var = tk.StringVar()
            ttk.Entry(form, textvariable=var, width=30).grid(row=i//2, column=(i%2)*2+1, padx=3, pady=2)
            self.vars[col] = var

        actions = ttk.Frame(self.root)
        actions.pack(fill="x", pady=5)

        ttk.Button(actions, text="Добавить", command=self.add_record).pack(side="left", padx=5)
        ttk.Button(actions, text="Загрузить из таблицы", command=self.load_selected_to_form).pack(side="left", padx=5)
        ttk.Button(actions, text="Обновить", command=self.update_record).pack(side="left", padx=5)
        ttk.Button(actions, text="Удалить", command=self.delete_record).pack(side="left", padx=5)

    def open_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            self.wb = openpyxl.load_workbook(path)
            self.sheet = self.wb[self.wb.sheetnames[0]]
            self.filepath = path
            self.load_from_sheet()
            messagebox.showinfo("Открыто", f"Файл: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def create_new(self):
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.sheet.title = "Sheet1"
        for i, col in enumerate(COLUMNS, start=1):
            self.sheet.cell(row=1, column=i, value=col)
        self.data = []
        self.refresh_table()
        self.filepath = None

    def save_file(self):
        if not self.filepath:
            return self.save_as()
        self._write_to_sheet()
        self.wb.save(self.filepath)
        messagebox.showinfo("OK", "Сохранено")

    def save_as(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if not path:
            return
        self._write_to_sheet()
        self.wb.save(path)
        self.filepath = path
        messagebox.showinfo("OK", "Файл сохранён")

    def load_from_sheet(self):
        self.data = []
        rows = list(self.sheet.iter_rows(values_only=True))
        if not rows:
            return
        headers = rows[0]
        for row in rows[1:]:
            record = {
                COLUMNS[i]: (str(row[i]) if i < len(row) and row[i] is not None else "")
                for i in range(len(COLUMNS))
            }
            self.data.append(record)
        self.refresh_table()

    def _write_to_sheet(self):
        self.sheet.delete_rows(1, self.sheet.max_row)
        for i, col in enumerate(COLUMNS, start=1):
            self.sheet.cell(row=1, column=i, value=col)
        for r, rec in enumerate(self.data, start=2):
            for c, col in enumerate(COLUMNS, start=1):
                self.sheet.cell(row=r, column=c, value=rec[col])

    def add_record(self):
        new_rec = {col: self.vars[col].get().strip() for col in COLUMNS}
        if not new_rec["Дата"]:
            new_rec["Дата"] = datetime.now().strftime(DATE_FORMAT)
        self.data.append(new_rec)
        self.refresh_table()

    def load_selected_to_form(self):
        selected = self.tree.selection()
        if not selected:
            return messagebox.showwarning("Выбор", "Выберите строку")
        index = int(selected[0])
        rec = self.data[index]
        for col in COLUMNS:
            self.vars[col].set(rec[col])

    def update_record(self):
        selected = self.tree.selection()
        if not selected:
            return messagebox.showwarning("Нет выбора", "Выберите строку")
        index = int(selected[0])
        updated_rec = {col: self.vars[col].get().strip() for col in COLUMNS}
        self.data[index] = updated_rec
        self.refresh_table()

    def delete_record(self):
        selected = self.tree.selection()
        if not selected:
            return
        index = int(selected[0])
        del self.data[index]
        self.refresh_table()

    def apply_filter(self):
        pattern = self.filter_var.get().strip().lower()
        if not pattern:
            return self.refresh_table()
        filtered_data = []
        for rec in self.data:
            if any(pattern in str(value).lower() for value in rec.values()):
                filtered_data.append(rec)
        self.refresh_table(filtered_data)

    def reset_filter(self):
        self.filter_var.set("")
        self.refresh_table()

    def refresh_table(self, items=None):
        for item in self.tree.get_children():
            self.tree.delete(item)
        data_source = items or self.data
        for i, rec in enumerate(data_source):
            values = [rec[col] for col in COLUMNS]
            self.tree.insert("", "end", iid=str(i), values=values)


if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1300x650")
    app = InventoryApp(root)
    root.mainloop()
